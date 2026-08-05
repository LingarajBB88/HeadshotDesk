"""
Turn an uploaded participant list into CSV text, whatever it arrived as.

Photographers get lists in whatever their client happens to use: a CSV
export, an Excel file, or a Numbers document from a Mac. Asking them to
convert first is the kind of small friction that gets a tool abandoned,
so we convert on our side and hand plain CSV to the existing importer.

Everything funnels into `to_csv_text`, so the import logic (validation,
dedupe, slot booking) has exactly one input format to reason about.
"""
from __future__ import annotations

import csv
import io
import logging
from pathlib import Path

from fastapi import HTTPException, status

logger = logging.getLogger(__name__)

SUPPORTED_SUFFIXES = {".csv", ".txt", ".tsv", ".xlsx", ".xlsm", ".numbers"}


def _rows_to_csv(rows: list[list[object]]) -> str:
    """Serialize a grid back to CSV, dropping trailing blank rows."""
    while rows and not any(str(c or "").strip() for c in rows[-1]):
        rows.pop()
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow(
            ["" if c is None else str(c).strip() for c in row]
        )
    return buf.getvalue()


def _xlsx_to_csv(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - dependency is declared
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel support isn't available on this server.",
        ) from e

    try:
        # read_only + data_only: we want values, not formulas, and we
        # never need styling.
        wb = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        sheet = wb.worksheets[0]
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    except Exception as e:  # noqa: BLE001
        logger.exception("Could not read xlsx upload")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "That Excel file couldn't be read. Try re-saving it as .xlsx "
                "or exporting to CSV."
            ),
        ) from e
    return _rows_to_csv(rows)


def _numbers_to_csv(content: bytes) -> str:
    try:
        from numbers_parser import Document
    except ImportError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Numbers files aren't supported on this server yet. In "
                "Numbers: File → Export To → CSV, then upload that."
            ),
        ) from e

    # numbers-parser needs a real path; the format is a package of files.
    import tempfile

    try:
        with tempfile.NamedTemporaryFile(suffix=".numbers") as tmp:
            tmp.write(content)
            tmp.flush()
            doc = Document(tmp.name)
            table = doc.sheets[0].tables[0]
            rows = [
                [cell.value for cell in row] for row in table.rows()
            ]
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        logger.exception("Could not read .numbers upload")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "That Numbers file couldn't be read. In Numbers: "
                "File → Export To → CSV, then upload that instead."
            ),
        ) from e
    return _rows_to_csv(rows)


def to_csv_text(filename: str, content: bytes) -> str:
    """Convert an uploaded participant list to CSV text.

    Raises a 400 with actionable wording for anything we can't read — the
    photographer is usually standing in front of a client when this fails.
    """
    suffix = Path(filename or "").suffix.lower()

    if suffix in {".xlsx", ".xlsm"}:
        return _xlsx_to_csv(content)
    if suffix == ".numbers":
        return _numbers_to_csv(content)
    if suffix in {".xls", ".ods", ".pages", ".pdf", ".docx"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"{suffix} files aren't supported. Save the list as CSV or "
                "Excel (.xlsx) and upload that."
            ),
        )

    # Plain text: CSV, TSV, or an unlabelled export. utf-8-sig strips the
    # BOM Excel writes; latin-1 is the fallback that never raises.
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Couldn't read that file. Try exporting it as CSV.",
    )
